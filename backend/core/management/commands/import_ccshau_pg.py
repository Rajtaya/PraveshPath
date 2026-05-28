import re
import openpyxl
from pathlib import Path
from django.core.management.base import BaseCommand
from core.models import (
    University, Course, UniversityCourse,
    EligibilityCriteria, AdmissionCycle,
)

EXCEL_PATH = Path(__file__).resolve().parents[4] / 'Data ' / 'Haryana_Universities_CCSHAU_PG_Course_Eligibility.xlsx'

STREAM_MAP = {
    'P301': 'agriculture', 'P302': 'agriculture', 'P303': 'agriculture',
    'P304': 'agriculture', 'P305': 'agriculture', 'P306': 'agriculture',
    'P307': 'agriculture', 'P308': 'agriculture', 'P309': 'agriculture',
    'P310': 'agriculture', 'P311': 'agriculture', 'P312': 'agriculture',
    'P313': 'agriculture', 'P314': 'agriculture',
    'P315': 'management', 'P316': 'management', 'P317': 'management',
    'P318': 'management', 'P319': 'management',
    'P320': 'engineering', 'P321': 'engineering', 'P322': 'engineering',
    'P323': 'engineering',
    'P335': 'science', 'P336': 'agriculture', 'P337': 'science',
    'P338': 'science', 'P339': 'science',
    'P340': 'science', 'P341': 'agriculture', 'P342': 'science',
    'P343': 'science', 'P344': 'science',
    'P345': 'agriculture', 'P346': 'agriculture', 'P347': 'agriculture',
    'P348': 'agriculture', 'P349': 'agriculture', 'P350': 'agriculture',
    'P351': 'agriculture', 'P352': 'agriculture',
    'P353': 'arts', 'P354': 'arts', 'P355': 'science',
    'P356': 'sports', 'P357': 'agriculture',
}

STATUS_MAP = {
    'notified-separately': 'upcoming',
    'closed': 'closed',
    'open': 'open',
}


def make_short_name(name):
    name = re.sub(r'\s*\(\d+\s*Year\)', '', name)
    name = re.sub(r'\s*\(\d+\s*Months?\)', '', name)
    short = name
    short = short.replace('Master in ', 'M')
    short = short.replace('PG Diploma in ', 'PGD ')
    short = re.sub(r'\s*-\s*(Dept\.|IBMA|COB).*', '', short)
    short = re.sub(r'\s*\(PGDC-\d+\)', '', short)
    short = re.sub(r'\s*\(.*?\)', '', short)
    short = short.strip()
    if len(short) > 30:
        short = short[:30].rstrip()
    return short


def read_sheet(wb, name):
    ws = wb[name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


class Command(BaseCommand):
    help = 'Import CCSHAU PG courses from Haryana_Universities_CCSHAU_PG_Course_Eligibility.xlsx'

    def handle(self, *args, **options):
        if not EXCEL_PATH.exists():
            self.stderr.write(f'Excel file not found: {EXCEL_PATH}')
            return

        wb = openpyxl.load_workbook(EXCEL_PATH)
        uni_data = read_sheet(wb, 'Universities')[0]
        courses_data = read_sheet(wb, 'Courses')
        cc_data = read_sheet(wb, 'College_Courses')
        elig_data = read_sheet(wb, 'Eligibility_Criteria')
        cycle_data = read_sheet(wb, 'Admission_Cycles')

        cc_by_id = {r['college_course_id']: r for r in cc_data}
        elig_by_id = {r['college_course_id']: r for r in elig_data}
        cycle_by_id = {r['college_course_id']: r for r in cycle_data}
        course_to_cc = {}
        for r in cc_data:
            course_to_cc[r['course_id']] = r['college_course_id']

        ccshau, created = University.objects.update_or_create(
            short_name='CCSHAU',
            defaults={
                'name': uni_data['name'],
                'university_type': uni_data['university_type'],
                'city': uni_data['city'],
                'district': uni_data['district'],
                'website': uni_data['website'] or '',
                'established_year': uni_data['established_year'],
                'is_active': True,
            }
        )
        action = 'Created' if created else 'Updated'
        self.stdout.write(f'{action} university: {ccshau}')

        course_created = 0
        uc_created = 0

        for cd in courses_data:
            cid = cd['course_id']
            ccid = course_to_cc.get(cid)
            if not ccid:
                self.stdout.write(f'  Skipping {cid} — no college_course mapping')
                continue

            cc = cc_by_id[ccid]
            elig = elig_by_id.get(ccid, {})
            cycle = cycle_by_id.get(ccid, {})

            stream = STREAM_MAP.get(cid, 'other')
            short_name = make_short_name(cd['name'])

            course, c_new = Course.objects.get_or_create(
                name=cd['name'],
                level='pg',
                defaults={
                    'short_name': short_name,
                    'stream': stream,
                    'duration_years': cd['duration_years'],
                }
            )
            if c_new:
                course_created += 1
            else:
                course.stream = stream
                course.duration_years = cd['duration_years']
                course.save()

            fee = cc.get('annual_fee')
            if fee == 'NULL' or fee is None:
                fee = None
            seats = cc.get('total_seats')
            if seats == 'NULL' or seats is None:
                seats = None

            uc, uc_new = UniversityCourse.objects.get_or_create(
                university=ccshau,
                course=course,
                defaults={
                    'total_seats': seats,
                    'annual_fee': fee,
                }
            )
            if not uc_new:
                uc.total_seats = seats
                uc.annual_fee = fee
                uc.save()
            else:
                uc_created += 1

            min_grad = elig.get('min_bachelor_percentage')
            if min_grad == 'NULL' or min_grad is None:
                min_grad = None

            req_subj = elig.get('required_subjects', '')
            if req_subj == '-' or req_subj is None:
                req_subj = ''

            note = elig.get('note', '') or ''

            EligibilityCriteria.objects.update_or_create(
                university_course=uc,
                defaults={
                    'min_graduation_percentage': min_grad,
                    'required_graduation': elig.get('required_bachelor_degree', ''),
                    'required_subjects': req_subj,
                    'entrance_exam': elig.get('entrance_exam', ''),
                    'domicile_required': str(elig.get('domicile_required', 'FALSE')).upper() == 'TRUE',
                }
            )

            raw_status = cycle.get('status', 'upcoming')
            status = STATUS_MAP.get(raw_status, 'upcoming')

            app_start = cycle.get('application_start')
            app_end = cycle.get('application_end')
            if app_start == 'NULL' or app_start is None:
                app_start = '2025-07-01'
            if app_end == 'NULL' or app_end is None:
                app_end = '2025-08-31'

            AdmissionCycle.objects.update_or_create(
                university_course=uc,
                academic_year=cycle.get('academic_year', '2025-26'),
                defaults={
                    'application_start': app_start,
                    'application_end': app_end,
                    'application_link': '',
                    'status': status,
                    'notes': f'{note}; Admission mode: {cycle.get("admission_mode", "")}',
                }
            )

        self.stdout.write(self.style.SUCCESS(
            f'\nCCSHAU PG import complete!\n'
            f'  New courses created: {course_created}\n'
            f'  University-courses created: {uc_created}\n'
            f'\n  DB totals:\n'
            f'    Universities: {University.objects.count()}\n'
            f'    Courses: {Course.objects.count()}\n'
            f'    University-courses: {UniversityCourse.objects.count()}\n'
        ))
