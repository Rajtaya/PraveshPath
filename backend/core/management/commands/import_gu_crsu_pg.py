import re
import openpyxl
from pathlib import Path
from django.core.management.base import BaseCommand
from core.models import (
    University, Course, UniversityCourse,
    EligibilityCriteria, AdmissionCycle,
)

EXCEL_PATH = Path(__file__).resolve().parents[4] / 'Data ' / 'Haryana_Universities_PG_Course_Eligibility.xlsx'

STREAM_MAP = {
    'medical': 'medical',
    'science-pcb': 'science',
    'science-pcm': 'science',
    'maths': 'science',
    'any': 'any',
}


def make_short_name(name):
    short = re.sub(r'\s*\(\d+\s*Year\)', '', name)
    short = re.sub(r'\s*\(.*?\)', '', short).strip()
    if len(short) > 30:
        short = short[:30].rstrip()
    return short


def read_sheet(wb, name):
    ws = wb[name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


class Command(BaseCommand):
    help = 'Import GU & CRSU PG courses from Haryana_Universities_PG_Course_Eligibility.xlsx'

    def handle(self, *args, **options):
        if not EXCEL_PATH.exists():
            self.stderr.write(f'Excel file not found: {EXCEL_PATH}')
            return

        wb = openpyxl.load_workbook(EXCEL_PATH)
        unis_data = read_sheet(wb, 'Universities')
        colleges_data = read_sheet(wb, 'Colleges')
        courses_data = read_sheet(wb, 'Courses')
        cc_data = read_sheet(wb, 'College_Courses')
        elig_data = read_sheet(wb, 'Eligibility_Criteria')

        college_to_uni = {c['college_id']: c['university_id'] for c in colleges_data}
        cc_by_id = {r['college_course_id']: r for r in cc_data}
        elig_by_id = {r['college_course_id']: r for r in elig_data}

        cc_by_course = {}
        for r in cc_data:
            cc_by_course.setdefault(r['course_id'], []).append(r)

        uni_objects = {}
        for ud in unis_data:
            uni, created = University.objects.update_or_create(
                short_name=ud['short_name'],
                defaults={
                    'name': ud['name'],
                    'university_type': ud['university_type'],
                    'city': ud['city'],
                    'district': ud['district'],
                    'website': ud['website'] or '',
                    'established_year': ud['established_year'],
                    'is_active': True,
                }
            )
            uni_objects[ud['university_id']] = uni
            self.stdout.write(f'{"Created" if created else "Updated"} university: {uni}')

        course_created = 0
        uc_created = 0

        for cd in courses_data:
            cid = cd['course_id']
            cc_list = cc_by_course.get(cid, [])
            if not cc_list:
                self.stdout.write(f'  Skipping {cid} — no college_course mapping')
                continue

            excel_stream = cd['stream'] or 'any'
            stream = STREAM_MAP.get(excel_stream, 'other')
            short_name = make_short_name(cd['name'])

            course, c_new = Course.objects.get_or_create(
                name=cd['name'],
                level='pg',
                defaults={
                    'short_name': short_name,
                    'stream': stream,
                    'duration_years': cd['duration_years'],
                }
            )
            if c_new:
                course_created += 1

            for cc in cc_list:
                ccid = cc['college_course_id']
                college_id = cc['college_id']
                uni_id = college_to_uni[college_id]
                uni = uni_objects[uni_id]

                fee = cc.get('annual_fee')
                if fee == 'NULL' or fee is None:
                    fee = None
                seats = cc.get('total_seats')
                if seats == 'NULL' or seats is None:
                    seats = None

                uc, uc_new = UniversityCourse.objects.get_or_create(
                    university=uni,
                    course=course,
                    defaults={
                        'total_seats': seats,
                        'annual_fee': fee,
                    }
                )
                if not uc_new:
                    uc.total_seats = seats
                    uc.annual_fee = fee
                    uc.save()
                else:
                    uc_created += 1

                elig = elig_by_id.get(ccid, {})
                min_grad = elig.get('min_bachelor_percentage')
                if min_grad == 'NULL' or min_grad is None:
                    min_grad = None

                req_subj = elig.get('required_subjects', '')
                if req_subj == '-' or req_subj is None:
                    req_subj = ''

                EligibilityCriteria.objects.update_or_create(
                    university_course=uc,
                    defaults={
                        'min_graduation_percentage': min_grad,
                        'required_graduation': elig.get('required_bachelor_degree', '') or '',
                        'required_subjects': req_subj,
                        'entrance_exam': elig.get('entrance_exam', '') or '',
                        'domicile_required': str(elig.get('domicile_required', 'FALSE')).upper() == 'TRUE',
                    }
                )

                AdmissionCycle.objects.update_or_create(
                    university_course=uc,
                    academic_year='2025-26',
                    defaults={
                        'application_start': '2025-07-01',
                        'application_end': '2025-08-31',
                        'application_link': '',
                        'status': 'upcoming',
                    }
                )

                self.stdout.write(f'  {uni.short_name}: {course.name} ({stream})')

        self.stdout.write(self.style.SUCCESS(
            f'\nGU & CRSU PG import complete!\n'
            f'  New courses created: {course_created}\n'
            f'  University-courses created: {uc_created}\n'
            f'\n  DB totals:\n'
            f'    Universities: {University.objects.count()}\n'
            f'    Courses: {Course.objects.count()}\n'
            f'    University-courses: {UniversityCourse.objects.count()}\n'
        ))
