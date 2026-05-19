import re
import openpyxl
from pathlib import Path
from django.core.management.base import BaseCommand
from core.models import (
    University, Course, UniversityCourse,
    EligibilityCriteria, AdmissionCycle,
)

EXCEL_PATH = Path(__file__).resolve().parents[4] / 'Data ' / 'Haryana_Universities_KUK_PG_Course_Eligibility.xlsx'

STREAM_MAP_BY_ID = {
    'P246': 'pharmacy',
}

STREAM_PATTERNS = [
    (r'^M\.Tech', 'engineering'),
    (r'^MBA', 'management'),
    (r'^LL\.M', 'law'),
    (r'^MCA\b|^Master of Computer Application', 'computer'),
    (r'^M\.Sc\.?\s*Computer', 'computer'),
    (r'^M\.Sc\.?\s*\(Graphics', 'design'),
    (r'^M\.Sc\.?\s*\(Printing', 'design'),
    (r'^M\.A\.?\s*\(Fine Arts\)', 'design'),
    (r'^Master of Fine Arts', 'design'),
    (r'^M\.A\.?\s*Music', 'design'),
    (r'^Master of Hotel', 'management'),
    (r'^Master of Tourism', 'management'),
    (r'Diploma.*Hospitality', 'management'),
    (r'^M\.Com', 'commerce'),
    (r'^M\.A\.?\s*Business Economics', 'commerce'),
    (r'^M\.A\.?\s*Education', 'education'),
    (r'^M\.Ed', 'education'),
    (r'^M\.P\.Ed|^Master of Physical Education', 'sports'),
    (r'^B\.P\.Ed|^Bachelor of Physical Education', 'sports'),
    (r'^M\.A\.?\s*Yoga', 'sports'),
    (r'Diploma.*Yoga', 'sports'),
    (r'^M\. Pharmacy', 'pharmacy'),
    (r'^M\.Sc', 'science'),
    (r'Diploma.*Floriculture', 'science'),
    (r'Diploma.*Health', 'science'),
    (r'^M\.A', 'arts'),
    (r'^M\.Lib', 'arts'),
    (r'^Master of Social Work', 'arts'),
    (r'^Certificate|Diploma.*Translation', 'arts'),
    (r'Diploma.*(Sanskrit|Buddhist|Archives|Women|Guidance|Counselling)', 'arts'),
    (r'Diploma', 'other'),
]


def resolve_stream(course_id, course_name, excel_stream):
    if course_id in STREAM_MAP_BY_ID:
        return STREAM_MAP_BY_ID[course_id]
    if excel_stream == 'science-pcb':
        return 'science'
    if excel_stream == 'science-pcm':
        return 'science'
    if excel_stream == 'maths':
        return 'science'
    if excel_stream == 'medical':
        return 'pharmacy'
    for pattern, stream in STREAM_PATTERNS:
        if re.search(pattern, course_name):
            return stream
    return 'other'


def make_short_name(name):
    name = re.sub(r'\s*\(\d+\s*Year\)', '', name)
    name = re.sub(r'\s*\(\d+\s*Months?\)', '', name)
    short = name
    short = short.replace('Master of ', 'M')
    short = short.replace('Bachelor of ', 'B')
    short = short.replace('Certificate Programme in ', 'Cert ')
    short = short.replace('PG Diploma in ', 'PGD ')
    short = short.replace('P.G. Diploma in ', 'PGD ')
    short = re.sub(r'\s*\(.*?\)', '', short)
    short = short.strip()
    if len(short) > 30:
        short = short[:30].rstrip()
    return short


def read_sheet(wb, name):
    ws = wb[name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


class Command(BaseCommand):
    help = 'Import KUK PG courses from Haryana_Universities_KUK_PG_Course_Eligibility.xlsx'

    def handle(self, *args, **options):
        if not EXCEL_PATH.exists():
            self.stderr.write(f'Excel file not found: {EXCEL_PATH}')
            return

        wb = openpyxl.load_workbook(EXCEL_PATH)
        courses_data = read_sheet(wb, 'Courses')
        cc_data = read_sheet(wb, 'College_Courses')
        elig_data = read_sheet(wb, 'Eligibility_Criteria')
        cycle_data = read_sheet(wb, 'Admission_Cycles')

        cc_by_id = {r['college_course_id']: r for r in cc_data}
        elig_by_id = {r['college_course_id']: r for r in elig_data}
        cycle_by_id = {r['college_course_id']: r for r in cycle_data}
        course_to_cc = {}
        for r in cc_data:
            course_to_cc[r['course_id']] = r['college_course_id']

        kuk = University.objects.get(short_name='KUK')

        course_created = 0
        uc_created = 0

        for cd in courses_data:
            cid = cd['course_id']
            ccid = course_to_cc.get(cid)
            if not ccid:
                self.stdout.write(f'  Skipping {cid} — no college_course mapping')
                continue

            cc = cc_by_id[ccid]
            elig = elig_by_id.get(ccid, {})
            cycle = cycle_by_id.get(ccid, {})

            stream = resolve_stream(cid, cd['name'], cd['stream'])
            short_name = make_short_name(cd['name'])

            course, c_new = Course.objects.get_or_create(
                name=cd['name'],
                level='pg',
                defaults={
                    'short_name': short_name,
                    'stream': stream,
                    'duration_years': cd['duration_years'],
                }
            )
            if c_new:
                course_created += 1
            else:
                course.stream = stream
                course.duration_years = cd['duration_years']
                course.save()

            fee = cc.get('annual_fee')
            if fee == 'NULL' or fee is None:
                fee = None

            uc, uc_new = UniversityCourse.objects.get_or_create(
                university=kuk,
                course=course,
                defaults={
                    'total_seats': cc.get('total_seats'),
                    'annual_fee': fee,
                }
            )
            if not uc_new:
                uc.total_seats = cc.get('total_seats')
                uc.annual_fee = fee
                uc.save()
            else:
                uc_created += 1

            min_grad = elig.get('min_bachelor_percentage')
            if min_grad == 'NULL' or min_grad is None:
                min_grad = None

            req_subj = elig.get('required_subjects', '')
            if req_subj == '-' or req_subj is None:
                req_subj = ''

            note = elig.get('note', '')
            if note is None:
                note = ''

            EligibilityCriteria.objects.update_or_create(
                university_course=uc,
                defaults={
                    'min_graduation_percentage': min_grad,
                    'required_graduation': elig.get('required_bachelor_degree', ''),
                    'required_subjects': req_subj,
                    'entrance_exam': elig.get('entrance_exam', ''),
                    'domicile_required': str(elig.get('domicile_required', 'FALSE')).upper() == 'TRUE',
                }
            )

            app_start = cycle.get('application_start', '2025-07-10')
            app_end = cycle.get('application_end', '2025-07-30')
            status = cycle.get('status', 'closed')

            AdmissionCycle.objects.update_or_create(
                university_course=uc,
                academic_year=cycle.get('academic_year', '2025-26'),
                defaults={
                    'application_start': app_start,
                    'application_end': app_end,
                    'application_link': 'https://iums.kuk.ac.in/anon_admissionHome.htm',
                    'status': status,
                    'notes': f'{note}; Admission mode: {cycle.get("admission_mode", "")}',
                }
            )

        self.stdout.write(self.style.SUCCESS(
            f'\nKUK PG import complete!\n'
            f'  New courses created: {course_created}\n'
            f'  University-courses created: {uc_created}\n'
            f'\n  DB totals:\n'
            f'    Universities: {University.objects.count()}\n'
            f'    Courses: {Course.objects.count()}\n'
            f'    University-courses: {UniversityCourse.objects.count()}\n'
            f'    UG courses: {Course.objects.filter(level="ug").count()}\n'
            f'    PG courses: {Course.objects.filter(level="pg").count()}\n'
        ))
