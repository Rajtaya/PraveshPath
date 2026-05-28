import re
import openpyxl
from datetime import datetime
from pathlib import Path
from django.core.management.base import BaseCommand
from core.models import (
    University, Course, UniversityCourse,
    EligibilityCriteria, AdmissionCycle,
)

EXCEL_PATH = Path(__file__).resolve().parents[4] / 'Data ' / 'GJUST_UG_Course_Eligibility.xlsx'

STATUS_MAP = {
    'closed': 'closed',
    'open': 'open',
    'unknown': 'upcoming',
    'notified-separately': 'upcoming',
}


def parse_seats(val):
    if val is None or val == 'NULL':
        return None
    s = str(val).strip()
    if '+' in s:
        try:
            return sum(int(x.strip()) for x in s.split('+'))
        except ValueError:
            return None
    try:
        return int(s)
    except ValueError:
        return None


def parse_date(val):
    if val is None or val == 'NULL':
        return None
    s = str(val).strip()
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def make_short_name(name):
    short = re.sub(r'\s*-\s*\d+\s*Year', '', name, flags=re.I)
    short = re.sub(r'\s*\(.*?\)', '', short).strip()
    if len(short) > 30:
        short = short[:30].rstrip()
    return short


def read_sheet(wb, name):
    ws = wb[name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


class Command(BaseCommand):
    help = 'Import GJUS&T UG courses from GJUST_UG_Course_Eligibility.xlsx'

    def handle(self, *args, **options):
        if not EXCEL_PATH.exists():
            self.stderr.write(f'Excel file not found: {EXCEL_PATH}')
            return

        wb = openpyxl.load_workbook(EXCEL_PATH)
        uni_data = read_sheet(wb, 'Universities')[0]
        courses_data = read_sheet(wb, 'Courses')
        cc_data = read_sheet(wb, 'College_Courses')
        elig_data = read_sheet(wb, 'Eligibility_Criteria')
        cycle_data = read_sheet(wb, 'Admission_Cycles')

        cc_by_id = {r['college_course_id']: r for r in cc_data}
        elig_by_id = {r['college_course_id']: r for r in elig_data}
        cycle_by_id = {r['college_course_id']: r for r in cycle_data}

        cc_by_course = {}
        for r in cc_data:
            cc_by_course.setdefault(r['course_id'], []).append(r)

        gjust, created = University.objects.update_or_create(
            short_name='GJUS&T',
            defaults={
                'name': uni_data['name'],
                'university_type': uni_data['university_type'],
                'city': uni_data['city'],
                'district': uni_data['district'],
                'website': uni_data['website'] or '',
                'established_year': uni_data['established_year'],
                'is_active': True,
            }
        )
        self.stdout.write(f'{"Created" if created else "Updated"} university: {gjust}')

        course_created = 0
        uc_created = 0

        for cd in courses_data:
            cid = cd['course_id']
            cc_list = cc_by_course.get(cid, [])
            if not cc_list:
                self.stdout.write(f'  Skipping {cid} — no college_course mapping')
                continue

            stream = cd['stream'] or 'any'
            short_name = make_short_name(cd['name'])

            course, c_new = Course.objects.get_or_create(
                name=cd['name'],
                level='ug',
                defaults={
                    'short_name': short_name,
                    'stream': stream,
                    'duration_years': cd['duration_years'],
                }
            )
            if c_new:
                course_created += 1

            for cc in cc_list:
                ccid = cc['college_course_id']

                fee = cc.get('annual_fee')
                if fee == 'NULL' or fee is None:
                    fee = None
                seats = parse_seats(cc.get('total_seats'))

                uc, uc_new = UniversityCourse.objects.get_or_create(
                    university=gjust,
                    course=course,
                    defaults={
                        'total_seats': seats,
                        'annual_fee': fee,
                    }
                )
                if not uc_new:
                    uc.total_seats = seats
                    uc.annual_fee = fee
                    uc.save()
                else:
                    uc_created += 1

                elig = elig_by_id.get(ccid, {})
                min_pct = elig.get('min_qualifying_percentage')
                if min_pct == 'NULL' or min_pct is None:
                    min_pct = None

                req_subj = elig.get('required_subjects', '')
                if req_subj == '-' or req_subj is None:
                    req_subj = ''

                EligibilityCriteria.objects.update_or_create(
                    university_course=uc,
                    defaults={
                        'min_12th_percentage': min_pct,
                        'required_subjects': req_subj,
                        'entrance_exam': elig.get('entrance_exam', '') or '',
                        'domicile_required': str(elig.get('domicile_required', 'FALSE')).upper() == 'TRUE',
                    }
                )

                cycle = cycle_by_id.get(ccid, {})
                raw_status = cycle.get('status', 'upcoming')
                status = STATUS_MAP.get(raw_status, 'upcoming')

                app_start = parse_date(cycle.get('application_start')) or '2026-05-04'
                app_end = parse_date(cycle.get('application_end')) or '2026-06-01'

                AdmissionCycle.objects.update_or_create(
                    university_course=uc,
                    academic_year=cycle.get('academic_year', '2026-27'),
                    defaults={
                        'application_start': app_start,
                        'application_end': app_end,
                        'application_link': '',
                        'status': status,
                        'notes': f'Admission mode: {cycle.get("admission_mode", "")}',
                    }
                )

                self.stdout.write(f'  {course.name} — seats:{seats}, fee:{fee}')

        self.stdout.write(self.style.SUCCESS(
            f'\nGJUS&T UG import complete!\n'
            f'  New courses created: {course_created}\n'
            f'  University-courses created: {uc_created}\n'
            f'\n  DB totals:\n'
            f'    Universities: {University.objects.count()}\n'
            f'    Courses: {Course.objects.count()}\n'
            f'    University-courses: {UniversityCourse.objects.count()}\n'
        ))
